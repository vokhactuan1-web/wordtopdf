"""
Excel to PDF Converter
Hỗ trợ Unicode đầy đủ cho tiếng Việt
"""
import os
import tempfile
import urllib.request
from pathlib import Path
from typing import Optional, List, Tuple

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER

from ..logging.logger_setup import get_logger
from ..io.file_handler import FileHandler

logger = get_logger(__name__)


class FontManager:
    """Quản lý font Unicode"""
    
    SYSTEM_FONTS = [
        'C:\\Windows\\Fonts\\arial.ttf',
        'C:\\Windows\\Fonts\\times.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/System/Library/Fonts/Supplemental/Times New Roman.ttf',
    ]
    
    @classmethod
    def get_unicode_font(cls) -> Optional[str]:
        """
        Tìm hoặc tải font Unicode hỗ trợ tiếng Việt
        
        Returns:
            Optional[str]: Đường dẫn font hoặc None
        """
        # Kiểm tra font hệ thống
        for font_path in cls.SYSTEM_FONTS:
            if os.path.exists(font_path):
                logger.info(f"Tìm thấy font: {font_path}")
                return font_path
        
        # Tải font từ internet
        return cls._download_font()
    
    @classmethod
    def _download_font(cls) -> Optional[str]:
        """Tải font DejaVu Sans"""
        try:
            temp_dir = tempfile.gettempdir()
            font_path = os.path.join(temp_dir, 'DejaVuSans.ttf')
            
            if os.path.exists(font_path):
                logger.info(f"Font cached: {font_path}")
                return font_path
            
            logger.info("Đang tải font DejaVu Sans...")
            url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
            
            urllib.request.urlretrieve(url, font_path)
            logger.info(f"Đã tải font: {font_path}")
            return font_path
            
        except Exception as e:
            logger.error(f"Không thể tải font: {e}")
            return None
    
    @classmethod
    def register_fonts(cls) -> Tuple[str, str]:
        """
        Đăng ký fonts với ReportLab
        
        Returns:
            Tuple[str, str]: (font_regular, font_bold)
        """
        font_path = cls.get_unicode_font()
        
        if font_path:
            try:
                pdfmetrics.registerFont(TTFont('UnicodeFont', font_path))
                pdfmetrics.registerFont(TTFont('UnicodeFont-Bold', font_path))
                logger.info("Đã đăng ký font Unicode")
                return 'UnicodeFont', 'UnicodeFont-Bold'
            except Exception as e:
                logger.error(f"Lỗi đăng ký font: {e}")
        
        logger.warning("Sử dụng Helvetica (có thể lỗi tiếng Việt)")
        return 'Helvetica', 'Helvetica-Bold'


class ExcelToPDFConverter:
    """Class chuyển đổi Excel sang PDF"""
    
    def __init__(self):
        self.font_regular, self.font_bold = FontManager.register_fonts()
    
    def convert(self, input_path: Path, output_path: Optional[Path] = None) -> Path:
        """
        Chuyển Excel sang PDF
        
        Args:
            input_path: Đường dẫn file Excel
            output_path: Đường dẫn file PDF output (tùy chọn)
            
        Returns:
            Path: Đường dẫn file PDF đã tạo
        """
        output_path = FileHandler.ensure_output_path(output_path, input_path, '.pdf')
        
        logger.info(f"Đang đọc Excel: {input_path.name}")
        wb = load_workbook(input_path, data_only=True)
        
        # Tạo PDF document
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=25,
            leftMargin=25,
            topMargin=25,
            bottomMargin=20
        )
        
        elements = self._process_workbook(wb)
        
        logger.info("Đang tạo PDF...")
        doc.build(elements)
        logger.info(f"Đã tạo PDF: {output_path}")
        
        return output_path
    
    def _process_workbook(self, wb) -> List:
        """Xử lý workbook và tạo elements cho PDF"""
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            fontName=self.font_bold,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        sheet_count = len(wb.sheetnames)
        logger.info(f"Tìm thấy {sheet_count} sheet(s)")
        
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            logger.info(f"Xử lý sheet {idx+1}/{sheet_count}: {sheet_name}")
            
            # Tiêu đề sheet
            elements.append(Paragraph(f"<b>{sheet_name}</b>", title_style))
            elements.append(Spacer(1, 0.15*inch))
            
            # Xử lý data
            table = self._create_table(ws)
            if table:
                elements.append(table)
            else:
                elements.append(Paragraph("<i>Sheet trống</i>", styles['Normal']))
            
            # Page break giữa các sheet
            if idx < len(wb.sheetnames) - 1:
                elements.append(PageBreak())
        
        return elements
    
    def _create_table(self, ws) -> Optional[Table]:
        """Tạo bảng từ worksheet"""
        data = []
        max_cols = 0
        row_count = 0
        
        # Thu thập dữ liệu
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx > 500:  # Giới hạn
                logger.warning("Giới hạn 500 dòng đầu tiên")
                break
            
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append('')
                else:
                    cell_str = str(cell)
                    if len(cell_str) > 120:
                        cell_str = cell_str[:117] + '...'
                    row_data.append(cell_str)
            
            if any(row_data):
                data.append(row_data)
                max_cols = max(max_cols, len(row_data))
                row_count += 1
        
        logger.info(f"  → {row_count} dòng, {max_cols} cột")
        
        if not data or max_cols == 0:
            return None
        
        # Cân bằng số cột
        for row in data:
            while len(row) < max_cols:
                row.append('')
        
        # Tính độ rộng cột
        col_widths = self._calculate_column_widths(data, max_cols)
        
        # Tạo bảng
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(self._get_table_style())
        
        return table
    
    def _calculate_column_widths(self, data: List[List[str]], max_cols: int) -> List[float]:
        """Tính toán độ rộng cột thông minh"""
        page_width = landscape(A4)[0] - 50
        col_widths = []
        
        for col_idx in range(max_cols):
            max_len = 0
            for row in data[:min(30, len(data))]:
                if col_idx < len(row):
                    cell_text = str(row[col_idx])
                    # Ký tự Unicode chiếm nhiều space hơn
                    ascii_count = sum(1 for c in cell_text if ord(c) < 128)
                    unicode_count = len(cell_text) - ascii_count
                    effective_len = ascii_count + unicode_count * 1.5
                    max_len = max(max_len, effective_len)
            
            width = min(max(max_len * 6, 40), 180)
            col_widths.append(width)
        
        # Scale nếu quá rộng
        total = sum(col_widths)
        if total > page_width:
            col_widths = [w * page_width / total for w in col_widths]
        
        return col_widths
    
    def _get_table_style(self) -> TableStyle:
        """Tạo style cho bảng"""
        return TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), self.font_regular),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#2980B9')),
            
            # Zebra striping
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#ECF0F1')]),
            
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])


# Hàm helper để sử dụng trực tiếp
def convert_excel_to_pdf(input_path: Path, output_path: Optional[Path] = None) -> Path:
    """
    Chuyển Excel sang PDF
    
    Args:
        input_path: Đường dẫn file Excel
        output_path: Đường dẫn file PDF (tùy chọn)
        
    Returns:
        Path: Đường dẫn file PDF
    """
    converter = ExcelToPDFConverter()
    return converter.convert(input_path, output_path)